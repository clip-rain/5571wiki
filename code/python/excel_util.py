import io
import itertools
from re import UNICODE
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# region re variables and constants
default_header_style = {
    'start_color': 'A6A3A3',
    'end_color': 'A6A3A3',
    'fill_type': 'solid'
}

col_style = {'header': {'start_color': '666666',
                        'end_color': '666666',
                        'fill_type': 'solid',
                        'width': '10'
                        },
             'data': {'start_color': '871123',
                      'end_color': '123432',
                      'fill_type': 'solid'
                      }
             }

sex_mapping = {
    1: "male", 2: "female"
}
# endregion


# region Object of Excel
class Field(object):

    # counter for field names, it is global so that it can be used by all fields
    _counter = itertools.count()

    def __init__(self, name, index=0, width=0, convert_func=str, col_style=None):
        """
        name: header name, the first row in excel. mandatory.
        index: column number. it can be added by metaclass by default.
        width: column width. it can be added by metaclass by default.
        convert_func: function to convert the value to the type of the field.
        col_style: column style. it can be added by metaclass by default.
        """
        # excel metaclass will use this order to get definition order of field in class attributes,
        self.header_name = name
        self.order = next(Field._counter)
        self.index = index
        # the width of the column in excel. if it is 0, it will be calculated by the length of the header name. keep in mind, this method of computing the width is not accurate.
        self.width = len(self.header_name) if width == 0 else width
        self.convert_func = convert_func
        self.col_style = col_style if col_style else {'header': {'with': self.width}}
        super(Field, self).__init__()


class ExcelMetaclass(type):

    def __new__(cls, name, bases, attrs):
        """
        dynamic create class with excel header name, and column index, and convert function, column width and so on.
        __col_dict__: key is index of field, value is the name of field.
        __header_dict__: key is index of field, values is the header name which set in Field definition.
        __header_width_dict__: key is index of field, values is the width of column.
        __convert_funcs__: key is index of field, values is the convert function which used for the Field.

        """
        # no need to modify class if it is ExcelModel.
        if name == 'ExcelModel':
            return type.__new__(cls, name, bases, attrs)
        field_items = [item for item in attrs.items()
                       if isinstance(item[1], Field)]
        if len(field_items) > 0:
            header_name_dict = dict()
            header_width_dict = dict()
            column_dict = dict()
            convert_func_dict = dict()
            data_style_dict = dict()
            slot = set()
            exist_index_set = set()
            sorted_items = sorted(field_items, key=lambda x: x[1].order)
            for k, v in sorted_items:
                if v.index:
                    exist_index_set.add(v.index)
            index = 1
            for k, v in sorted_items:
                # k is the name of field, v is the field object.
                if not v.index:
                    # add a auto increment index when without index setting,
                    while index in exist_index_set:
                        index += 1
                    current_index = index
                    index += 1
                else:
                    current_index = v.index
                # set field name for debug.
                v.field_name = k
                header_name_dict[current_index] = v.header_name
                header_width_dict[current_index] = v.width
                column_dict[current_index] = k
                convert_func_dict[current_index] = v.convert_func
                data_style_dict[current_index] = v.col_style
                slot.add(k)
            attrs['__col_dict__'] = column_dict
            attrs['__header_dict__'] = header_name_dict
            attrs['__data_style_dict__'] = data_style_dict
            attrs['__header_width_dict__'] = header_width_dict
            attrs['__convert_funcs__'] = convert_func_dict
            attrs['__data_slot__'] = slot
        return type.__new__(cls, name, bases, attrs)


class ExcelModel(metaclass=ExcelMetaclass):

    converted_map = {
    }

    def __init__(self, **kw):
        self.__dict__['__data_dict__'] = dict(**kw)

    def __getattr__(self, key):
        try:
            return self.__dict__['__data_dict__'][key]
        except KeyError:
            if key in self.__data_slot__:
                return None
            raise AttributeError(r"'Model' object has no attribute '%s'" % key)

    def __setattr__(self, key, value):
        self.__dict__['__data_dict__'][key] = value

    def update(self, **kw):
        self.__dict__['__data_dict__'].update(**kw)

    def get(self, k):
        return self.__getattr__(k)

    def to_row_data(self):
        """
        convert to data.
        """
        d = dict()
        for col_index, col_name in self.__col_dict__.items():
            n = self.get(col_name)
            if n is None:
                cell_value = '-'
            else:
                if col_name in self.converted_map:
                    cell_value = self.converted_map[col_name](
                        self.get(col_name))
                else:
                    try:
                        func = self.__convert_funcs__.get(col_index, str)
                        cell_value = func(self.get(col_name))
                    except:
                        cell_value = str(self.get(col_name))
            # to avoid the illegal characters in excel.
            if isinstance(cell_value, str):
                cell_value = ILLEGAL_CHARACTERS_RE.sub(r'', cell_value)
            d[col_index] = cell_value
        return d

    def get_header_dict(self):
        return self.__header_dict__

    def get_header_width_dict(self):
        return self.__header_width_dict__

    def get_data_style_dict(self):
        return self.__data_style_dict__


# endregion


class RainExcel(object):
    """
    not thread safe.
    """
    suffix = '.xlsx'
    content_type = 'application/vnd.ms-excel'

    def __init__(self):
        self.workbook = Workbook()
        self.sheet_count = 0
        self.default_sheet = self.workbook.active
        super(RainExcel, self).__init__()

    def add_sheet(self, sheet_name, row_iterator):
        """
        add sheet to pas excel. if this sheet existed before, just append the rows to the end of the sheet.
        """
        has_header = True
        if sheet_name in self.workbook.sheetnames:
            ws = self.workbook.get_sheet_by_name(sheet_name)
        else:
            ws = self.workbook.create_sheet(sheet_name)
            self.sheet_count += 1
            has_header = False
        for row in row_iterator:
            # need to add header if it does have header.
            if not has_header:
                self._add_header(ws, row)
                has_header = True
            ws.append(row.to_row_data())

    def save(self, file_name):
        """
        save the excel to file.
        """
        if self.sheet_count != 0:
            self.workbook.remove(self.default_sheet)
        self.workbook.save(file_name)

    def get_excel_data(self):
        """
        get excel data as buffer which can be put into http response.
        """
        # remove the default empty sheet.
        if self.sheet_count != 0:
            self.workbook.remove(self.default_sheet)
        # workbook use zipfile to archive the data, so the parameter of save would be a filename or a file-like object.
        buf = io.StringIO()
        self.workbook.save(buf)
        return buf.getvalue()

    @classmethod
    def _add_header(cls, ws, row):
        """
        add header for a new work sheet.
        ws: the work sheet
        row: origin data to insert into excel.
        """
        ws.append(row.get_header_dict())
        cls._set_header_style(ws, row)

    @classmethod
    def _set_header_style(cls, ws, row):
        """
        set the style of header, e.g. width, colour.
        ws: the work sheet.
        row: origin data to insert into excel.
        """
        data_style = row.get_data_style_dict()
        dims = {}
        header_line = next(ws.rows)
        # change to use style dict instead of width dict.
        # header_width_dict = row.get_header_width_dict()
        data_style = row.get_data_style_dict()
        for cell in header_line:
            if cell.value:
                header_style_dict = data_style.get(
                    cell.col_idx, {}).get('header', {})
                header_cell_color_style = dict(start_color=header_style_dict.get('start_color', default_header_style.get('start_color')),
                                               end_color=header_style_dict.get('end_color',  default_header_style.get('end_color')),
                                               fill_type=header_style_dict.get('fill_type',  default_header_style.get('fill_type')))
                header_cell_with = header_style_dict.get('width', len(str(cell.value)))
                header_fill = PatternFill(**header_cell_color_style)
                cell.fill = header_fill
                # use the dims to set the width of the header.
                ws.column_dimensions[cell.column_letter].width = header_cell_with


    @classmethod
    def _set_data_style(cls, ws, row):
        # todo set the style of data.
        pass


class TestExcelA(ExcelModel):
    # use col_style to set the style of the column.
    name = Field("Name", col_style=col_style)
    # use index to appoint the position of the column.
    age = Field("Age", index=3)
    # use converter to convert the value of the column.
    sex = Field("Sex", convert_func=lambda x: sex_mapping.get(x, 'Unknown'))


class TestExcelB(ExcelModel):
    name = Field("Name")
    age = Field("Age")
    card_number = Field("Card Number")


# def get_excel_http_response(pas_excel, filename):
# 	"""
# 	get http response for excel
# 	"""
# 	response = HttpResponse(pas_excel.get_excel_data(), content_type=pas_excel.content_type)
# 	response['Content-Disposition'] = 'attachment; filename="%s"' % (filename + pas_excel.suffix)
# 	return response


if __name__ == "__main__":
    self_excel = RainExcel()
    sheet1_data = [
        {'name': 'test1', 'age': '12', 'sex': 1},
        {'name': 'test2', 'age': '13', 'sex': 2}
    ]
    sheet2_data = [
        {'name': 'test1', 'age': '12', 'card_number': '5432'},
        {'name': 'test2', 'age': '13', 'card_number': '1234'}
    ]
    sheet1_obj_data = [TestExcelA(**row) for row in sheet1_data]
    sheet2_obj_data = [TestExcelB(**row) for row in sheet2_data]
    self_excel.add_sheet('ASheet', sheet1_obj_data)
    self_excel.add_sheet('BSheet', sheet2_obj_data)
    # current path is the root path of this project.
    self_excel.save("./code/python/data/test.xlsx")
